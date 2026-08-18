"""
Build the sub-agent clarifying-questions xlsx.

Format mirrors the reference (Spec-Pipeline Decision Matrix):
  A: ID
  B: Topic
  C: Decision / Question
  D: Options
  E: Recommendation
  F: Your choice (pre-filled with default — overwrite to change)
  G: Rationale
  H: Impact if changed

18 questions across 7 decision areas. Each question has 2-4 options.
Pre-fills column F with the recommendation. The user overwrites F to
change the answer; column G / H explain why the recommendation stands
and what shifts if the user picks a different option.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\dev\pi-loop\wt-pi-subagent\docs\PRD\sub-agent-questions.xlsx"

# ---------------------------------------------------------------------------
# Questions
# ---------------------------------------------------------------------------
# Each entry: (id, topic, question, options, recommendation, your_choice,
#              rationale, impact)
#
# Options are written as "A) ...\nB) ...\nC) ..." strings (the reference
# uses \n as a row separator inside the cell; openpyxl preserves it).

QUESTIONS = [
    # -------- Group 1: Defaults & scope -----------------------------------
    (
        "D1",
        "Default isolation mode",
        "When a user creates a new loop with /loop, what is the default value of the new `isolation` field?",
        "A) `in-process` — current v2.x behaviour; the wake is a turn in the parent session\n"
        "B) `sub-agent` — new default; each fire spawns a child\n"
        "C) `ask` — prompt the user the first time, remember the answer per-project",
        "A — in-process",
        "A — in-process",
        "Preserves backwards compatibility for every existing loop. New users who want sub-agent mode opt in per-loop (or set `subAgent.defaultIsolation` in settings). A 'ask' UX adds a step to a high-frequency command.",
        "B would surprise every existing user; C adds friction to a fast command. Default is safe; the per-loop `isolation` field is always available.",
    ),
    (
        "D2",
        "Cross-platform day 1",
        "Which platforms must sub-agent mode work on at v2.5.0 release?",
        "A) macOS + Linux only; Windows is a follow-up release\n"
        "B) macOS + Linux + Windows (all three day 1)",
        "B — all three platforms",
        "B — all three platforms",
        "You're on Windows. The architecture (child_process.spawn, file-backed result store, atomic writes) is already cross-platform; the only Windows-specific concern is the `pi` binary path resolution. Building it for Windows day 1 is the same work as building it for Unix; deferring would just ship a half-feature.",
        "A is faster to ship but leaves the primary user (you) without the feature. Pick A only if there's a hard reason (e.g. a Windows-only bug we cannot solve).",
    ),
    (
        "D3",
        "Default per-iteration wall-clock timeout",
        "What is the default `subAgent.iterationTimeoutMs`? The cap is 24h; this is what every new loop inherits unless the user overrides it.",
        "A) 5 min (300,000 ms) — tight; matches quick polling loops\n"
        "B) 10 min (600,000 ms) — matches the PRD's draft default\n"
        "C) 30 min (1,800,000 ms) — accommodates moderate digests\n"
        "D) 1 hour (3,600,000 ms) — accommodates batch / research jobs",
        "B — 10 min",
        "B — 10 min",
        "10 min is a sweet spot: long enough for a meaningful chunk of work (a digest, a small audit, a research brief), short enough that a runaway child is killed before it costs serious money. Users with longer jobs override per-loop. The /loop-settings TUI cycles 5min → 10min → 30min → 1h → 6h → 24h → 5min.",
        "A kills legitimate 10-min research jobs too early; C/D let a single bad iteration burn 5–10x the intended budget before it's killed.",
    ),
    (
        "D4",
        "Default active-iteration concurrency cap",
        "How many sub-agent iterations may be in flight at once across all sub-agent loops in one session?",
        "A) 1 — serial; one iteration at a time, others queue\n"
        "B) 2 — light concurrency, low CPU pressure\n"
        "C) 4 — matches the PRD's draft default; sensible for a 4–8 core machine\n"
        "D) 8 — for power users with 16+ core machines",
        "C — 4",
        "C — 4",
        "4 is the right default for a typical dev machine. 1 wastes the trigger system's ability to fire overlapping loops; 8+ saturates CPU and starves the parent. The cap is per-session (so two terminals in the same project are independent), and the user can raise it in /loop-settings if they have the cores.",
        "A is too conservative — a 5-min cron loop with a 1-min digest would never overlap with itself anyway, but a 5-min cron and a 10-min event loop would. C lets them overlap without fighting for CPU.",
    ),
    # -------- Group 2: Result delivery -----------------------------------
    (
        "D5",
        "Result preview length and format",
        "The one-line summary that surfaces to the parent after each sub-agent iteration — how long and what does it contain?",
        "A) 200 chars, the same for all priorities; the user reads the file for full content\n"
        "B) 200 chars for normal/urgent/defer; 1,000 chars for critical (so the parent can act without opening the file)\n"
        "C) Fully configurable per-loop; no default override",
        "B — tiered by priority",
        "B — tiered by priority",
        "Critical loops exist exactly because the parent needs to act fast. Letting critical previews be longer (up to 1 KiB) means the parent can react in one beat; non-critical previews stay terse to protect the parent context. C is too much knob for the common case.",
        "A forces a 'critical' loop's parent to open the file before acting, defeating the point. C is fine as a v3 add-on but not day 1.",
    ),
    (
        "D6",
        "Critical-priority interrupt semantics",
        "When a `priority: critical` sub-agent loop completes, may it interrupt the parent's currently-running turn, or does it always wait for `agent_end`?",
        "A) Always wait for `agent_end` — never interrupt\n"
        "B) Allow interrupt only if the parent's current turn is not itself a loop wake (avoid interrupting a wake with another wake)\n"
        "C) Always interrupt — the user can mute critical loops in /loop-settings if they don't want the interruption",
        "B — interrupt only outside loop wakes",
        "B — interrupt only outside loop wakes",
        "Critical exists for security / safety / data-loss reasons. Interrupting the parent's user-driven turn is fine; interrupting a wake that is itself processing another critical is a re-entrancy footgun. B is the conservative default; users who want pure C can set `subAgent.criticalInterruptsAll: true` later.",
        "A makes critical de facto equal to urgent; C is the footgun. B is the safe default that can be loosened.",
    ),
    # -------- Group 3: Security & safety ---------------------------------
    (
        "D7",
        "Tool allowlist default",
        "When a sub-agent loop has no explicit `subAgent.tools` field, what tools does the child get?",
        "A) Full surface (read, grep, find, ls, edit, write, bash with 30s timeout)\n"
        "B) Read-only by default (read, grep, find, ls); the user must opt in to write/edit/bash\n"
        "C) No tools by default; the user must specify a list",
        "B — read-only by default",
        "B — read-only by default",
        "Sub-agent loops are typically digests, audits, and read-mostly work. A read-only default is the safe one. Users who want to write (a sub-agent that files a PR, a sub-agent that updates a doc) opt in explicitly with `subAgent.tools: [read, grep, find, edit, write]`. The cyclic field form in /loop-edit prompts for confirmation when a non-readonly allowlist is selected.",
        "A is the most flexible but the most dangerous — a sub-agent loop with a 5-min cron could quietly rewrite files every 5 min. C is too restrictive — most sub-agent loops want at least read.",
    ),
    (
        "D8",
        "Recursion guard",
        "How do we ensure a sub-agent child does not spawn its own sub-agents (a loop inside a loop)?",
        "A) Child starts with `--no-extensions` — cannot load `pi-loop` or `pi-subagents`; safe by default\n"
        "B) A — plus deny the `subagent` tool even if the child agent config includes it\n"
        "C) Allow nesting up to depth 2 with explicit user opt-in",
        "B — no-extensions + deny subagent",
        "B — no-extensions + deny subagent",
        "A is enough for the v2.5.0 case because the child does not load pi-subagents and therefore does not have the `subagent` tool. B is a belt-and-braces guarantee for the case where a future change accidentally re-enables extensions. C introduces a recursion budget and depth-tracking code that has no day-1 use case.",
        "A is the minimum; B is the minimum plus a safety net. C is a v3 feature for users who genuinely want nested sub-agents.",
    ),
    # -------- Group 4: Lifecycle & data ----------------------------------
    (
        "D9",
        "Result retention",
        "How long do sub-agent result files (`.pi/loops/sub-agent-results/<loopId>/iter-<N>/`) live on disk?",
        "A) 7 days, then auto-pruned\n"
        "B) 30 days, then auto-pruned\n"
        "C) Last 50 iterations per loop, oldest auto-pruned\n"
        "D) Unlimited; the user is responsible for cleanup",
        "C — last 50 iterations",
        "C — last 50 iterations",
        "Result files are small (1–10 KiB each) but accumulate at the trigger rate. 50 iterations of a 5-min loop is ~4 hours of history; 50 iterations of a weekly loop is ~1 year. The cap protects disk without losing the recent past. The user can override per-loop (`subAgent.retainIterations: 200`) or per-session (`subAgent.globalRetainIterations: 500`).",
        "A/B are time-based and align poorly with trigger rates; D is a footgun. C is bounded and intuitive.",
    ),
    (
        "D10",
        "Loop scope interaction",
        "A loop with `loopScope: 'project'` and `isolation: 'sub-agent'` — where do the result files live, and are they shared across terminals?",
        "A) Under the project root (`.pi/loops/sub-agent-results/...`) — shared across all terminals in this repo, but each terminal's `subAgentRunWatcher` only watches results for loops it has bound\n"
        "B) Per-session (`.pi/loops-<sessionId>/sub-agent-results/...`) — strictly local to the terminal that created the loop\n"
        "C) Both: a project-wide index pointing at per-session dirs, for cross-terminal visibility without shared file writes",
        "A — project root",
        "A — project root",
        "Matches the v2.x loop-scope invariant: project-scope = shared, session-scope = per-terminal. The `bindings-<sessionId>.json` file (per ADR-006) already enforces that only bound terminals fire the loop; the same gate keeps each terminal's watcher from doing work for unbound loops. C is clever but introduces a sync problem the user did not ask us to solve.",
        "B breaks cross-terminal visibility for a feature explicitly designed to handle long-running loops (which are most useful when they survive terminal swap). C is a v3+ concern.",
    ),
    # -------- Group 5: Integration ---------------------------------------
    (
        "D11",
        "pi-subagents background-work provider bridge",
        "If `pi-subagents` is installed, should `pi-loop` automatically register itself as a background-work provider so its iterations show in `pi-subagents`'s `/subagents-fleet`?",
        "A) Auto-register on first sub-agent spawn (lazy)\n"
        "B) Opt-in via `subAgent.registerBackgroundWorkProvider: true` in settings\n"
        "C) Never register; sub-agent iterations are visible only via the pi-loop TUI",
        "A — auto-register on first spawn",
        "A — auto-register on first spawn",
        "The integration is lazy and feature-detected (it checks for `Symbol.for('pi-subagents.background-work.v1')` at runtime). If the symbol is absent, no-op. If the user does not want the integration, they can set `subAgent.registerBackgroundWorkProvider: false` after first spawn. The default 'auto' maximises interop without forcing the user to know about it.",
        "B makes the user discover a setting they did not know existed; C is silent and rude when both packages are installed.",
    ),
    (
        "D12",
        "Capability ceiling auto-read",
        "If `pi-subagents` has registered a capability ceiling via `registerSubagentCapabilityCeiling`, should `pi-loop` honour it automatically?",
        "A) Yes — read the ceiling at spawn time and intersect with the loop's tool allowlist\n"
        "B) Yes, but log a warning the first time so the user knows it's happening\n"
        "C) No — pi-loop's sub-agent mode is independent; the user must enforce the ceiling in their loop config",
        "B — read with warning",
        "B — read with warning",
        "The capability ceiling is a defence-in-depth feature the user has explicitly turned on in `pi-subagents`. Honoring it is the right thing to do; the first-time warning makes the intersection visible (otherwise a user might wonder why their loop suddenly has fewer tools). A is silent and confusing; C makes the two extensions fight.",
        "A is correct in spirit but the silent intersection will look like a bug to the user. B is A + a log line that points to the right setting.",
    ),
    (
        "D13",
        "Cost ceiling backstop",
        "Beyond per-loop `maxTokens`, do we need a global `subAgent.costCeilingUsd` setting that hard-stops all sub-agent activity when the session's cumulative cost hits a number?",
        "A) No — per-loop `maxTokens` is enough\n"
        "B) Yes — global cost ceiling, default $5 / session, user-configurable in /loop-settings",
        "B — global cost ceiling",
        "B — global cost ceiling",
        "A is true for the well-configured user. B exists for the user who has 10 sub-agent loops each with `maxTokens: 1_000_000` and does not want a runaway to drain the wallet. A $5 default is roughly 200k tokens of sonnet — enough for a week of typical sub-agent work, low enough to catch a runaway. The setting is per-session (per-terminal), not global, so two terminals in the same project are independent.",
        "A ships a feature that will eventually burn a user. B is cheap to add (one number in cost-tracker) and prevents the worst-case support ticket.",
    ),
    # -------- Group 6: UX ------------------------------------------------
    (
        "D14",
        "FleetView panel visibility",
        "When does the new 'active sub-agents' panel appear?",
        "A) Always, even when no iterations are running (placeholder text: 'No active sub-agent loops')\n"
        "B) Only when at least one iteration is in `running` or `starting` state\n"
        "C) Always for sub-agent-loop users; hidden for users who have no sub-agent loops",
        "B — only when active",
        "B — only when active",
        "The panel is informational; an empty panel is noise. Showing it only when there is something to see matches the existing monitor and task widgets' behaviour. A wastes a row of the editor; C is harder to maintain (need to track 'has the user ever made a sub-agent loop?').",
        "A is too noisy; C requires per-user state we don't otherwise track. B is the right default.",
    ),
    (
        "D15",
        "LoopInspect as a tool",
        "Should `LoopInspect` be a tool (callable by the agent) or a slash command (callable only by the user)?",
        "A) Tool only\n"
        "B) Slash command only\n"
        "C) Both — same handler, two surfaces",
        "C — both",
        "C — both",
        "The agent uses the tool to inspect its own sub-agent runs and reason about them; the user uses the slash command to inspect interactively. Same handler, two surfaces. A excludes the user; B excludes the agent. C is the smallest surface that covers both.",
        "A breaks the user; B breaks the agent's introspective loops. C is correct.",
    ),
    # -------- Group 7: Naming & phasing ----------------------------------
    (
        "D16",
        "Field name",
        "The new per-loop field that switches execution mode. The PRD uses `isolation`. Other candidates?",
        "A) `isolation` — what the PRD currently uses; emphasises 'the work is isolated from the parent'\n"
        "B) `executionMode` — explicit 'where the work runs'\n"
        "C) `worker` — short; borrows from the 'subagent' / 'worker' agent vocabulary\n"
        "D) `runner` — short; borrows from `pi-subagents` runner types",
        "A — `isolation`",
        "A — `isolation`",
        "`isolation` is consistent with the existing `loopScope` field (which describes WHERE the loop's state lives); `isolation` describes HOW the iteration runs. `executionMode` is more explicit but verbose; `worker` collides with the agent name `worker` in `pi-subagents`; `runner` is too generic and the term 'runner' is already overloaded in `pi-subagents` (runner.type).",
        "B is the next-best alternative if `isolation` feels too abstract to the user. C/D create terminology collisions.",
    ),
    (
        "D17",
        "Phase 1 scope (v2.5.0)",
        "What ships in v2.5.0 vs v2.5.1?",
        "A) v2.5.0: settings + types + no-op sub-agent path (the field is persisted, but no spawn). v2.5.1: actual spawn + result + cost + TUI\n"
        "B) v2.5.0: everything (settings, spawn, result, cost, TUI) in one release\n"
        "C) v2.5.0: settings + types + spawn + result (no TUI, no fleet view). v2.5.1: TUI + fleet view + pi-subagents bridge",
        "C — v2.5.0 = backend, v2.5.1 = frontend",
        "C — v2.5.0 = backend, v2.5.1 = frontend",
        "Two minor releases lets us land the high-risk piece (the spawn / parent-restart reconciliation) in v2.5.0 and the user-facing TUI in v2.5.1 once the backend is stable. A is too cautious; B bundles the riskier and the less-risky work into one release with no rollback granularity.",
        "A makes the feature invisible to the user for a full release cycle. B is fine if the user wants a single release. C is the safest split.",
    ),
    (
        "D18",
        "Cross-machine visibility (P3 persona)",
        "A user running sub-agent loops on a server and inspecting from a desktop. Address in v2.5.x or defer?",
        "A) Defer to v3 — the user did not ask for this; v2.5.x is single-machine\n"
        "B) Address in v2.5.1 as a stretch — expose the FleetView state file over SSH / Tailscale\n"
        "C) Address in v2.5.0 — build the cross-machine path as a first-class concern",
        "A — defer to v3",
        "A — defer to v3",
        "You asked for loops that survive parent restarts. That is local-process restart, not cross-machine. Cross-machine adds SSH / Tailscale / file-sync concerns that are out of scope. A keeps v2.5.x focused. B is a stretch we can take if the core ships early; C is a non-trivial scope expansion that would delay the single-machine feature.",
        "B/C are fine if cross-machine is a real near-term need. Default to A and revisit.",
    ),
]

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Decision Matrix"

headers = [
    "ID",
    "Topic",
    "Decision / Question",
    "Options",
    "Recommendation",
    "Your choice (pre-filled with default — overwrite to change)",
    "Rationale",
    "Impact if changed",
]
ws.append(headers)

# Header style
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, _ in enumerate(headers, 1):
    cell = ws.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Body style
body_font = Font(name="Calibri", size=10, color="000000")
input_font = Font(name="Calibri", size=10, color="0000FF")  # blue = the cell the user overwrites
wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
recommend_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow = the recommended default
for q in QUESTIONS:
    qid, topic, question, options, recommend, your_choice, rationale, impact = q
    row = [qid, topic, question, options, recommend, your_choice, rationale, impact]
    ws.append(row)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(r, c)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = border
    # The 'Your choice' column is the one the user types into; highlight.
    ws.cell(r, 6).font = input_font
    # The recommendation cell is the suggested default; tint to draw the eye.
    ws.cell(r, 5).fill = recommend_fill

# Column widths
widths = {
    1: 6,    # ID
    2: 26,   # Topic
    3: 50,   # Decision / Question
    4: 50,   # Options
    5: 28,   # Recommendation
    6: 50,   # Your choice
    7: 50,   # Rationale
    8: 50,   # Impact
}
for c, w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# Row heights — tall enough for the wrapped question + options
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 180

# Freeze the header row
ws.freeze_panes = "A2"

# Tab color
ws.sheet_properties.tabColor = "305496"

# ---------------------------------------------------------------------------
# A second sheet: legend & how to answer
# ---------------------------------------------------------------------------
legend = wb.create_sheet("How to answer")
legend["A1"] = "Sub-agent PRD — Decision Matrix"
legend["A1"].font = Font(name="Calibri", size=14, bold=True, color="305496")
legend["A3"] = "What this is"
legend["A3"].font = Font(name="Calibri", size=11, bold=True)
legend["B3"] = ("18 questions for the sub-agent execution feature in pi-loop (see "
                "docs/PRD/sub-agent.md). Each question has 2–4 options, a recommendation "
                "(col E), and a pre-filled default answer (col F, blue).")
legend["B3"].alignment = wrap
legend["A5"] = "How to answer"
legend["A5"].font = Font(name="Calibri", size=11, bold=True)
legend["B5"] = ("1. Open the 'Decision Matrix' sheet.\n"
                "2. For each row, type your choice into column F (the blue one). "
                "   - If you take the recommendation, type the option letter (A, B, C, D) or the short form.\n"
                "   - If you want a hybrid or a custom answer, write it out in plain English.\n"
                "3. Add a free-text note at the end of the cell if you have a constraint the options do not cover.\n"
                "4. Save the file. Send it back; I will review and decide whether more questions are needed.\n"
                "5. If we agree on everything, I start Phase 1 (v2.5.0) implementation.")
legend["B5"].alignment = wrap
legend["A10"] = "Question groups"
legend["A10"].font = Font(name="Calibri", size=11, bold=True)
legend["B10"] = ("D1–D4   Defaults & scope\n"
                 "D5–D6   Result delivery\n"
                 "D7–D8   Security & safety\n"
                 "D9–D10  Lifecycle & data\n"
                 "D11–D13 Integration (pi-subagents, capability ceiling, cost ceiling)\n"
                 "D14–D15 UX (panel visibility, LoopInspect as tool)\n"
                 "D16–D18 Naming & phasing (field name, release split, cross-machine)")
legend["B10"].alignment = wrap
legend["A15"] = "Columns"
legend["A15"].font = Font(name="Calibri", size=11, bold=True)
legend["B15"] = ("A  ID            D1, D2, …\n"
                 "B  Topic         short label\n"
                 "C  Question      the actual decision to make\n"
                 "D  Options       A, B, C, D — what you can pick\n"
                 "E  Recommend     the suggested default (yellow tint)\n"
                 "F  Your choice   YOUR answer — pre-filled with the recommendation; overwrite\n"
                 "G  Rationale     why we recommend the default\n"
                 "H  Impact        what changes if you pick a different option")
legend["B15"].alignment = wrap
legend["A22"] = "Notes"
legend["A22"].font = Font(name="Calibri", size=11, bold=True)
legend["B22"] = ("All recommendations assume a Windows-first user on a 4–8 core dev machine. "
                 "If you are on a different setup, please say so in the relevant row.\n\n"
                 "The PRD draft is at docs/PRD/sub-agent.md (12k words). It is safe to skim §1 "
                 "(Executive summary) and §15 (Open questions) before answering; the rest is "
                 "implementation detail you do not need to read to make these decisions.")
legend["B22"].alignment = wrap
legend.column_dimensions["A"].width = 22
legend.column_dimensions["B"].width = 100
for r in (3, 5, 10, 15, 22):
    legend.row_dimensions[r].height = 110
legend.row_dimensions[1].height = 24

# Save
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  sheets: {wb.sheetnames}")
print(f"  questions: {len(QUESTIONS)}")
