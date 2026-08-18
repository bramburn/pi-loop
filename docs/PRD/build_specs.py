"""
Build the sub-agent feature spec xlsx — v2 (dict-based, no missing fields).

Each spec is a list of dicts with keys: id, title, what, how, accept, fail, test, loc.
The 8-tuple is built from the dict at write time, so a missing key is impossible.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\dev\pi-loop\wt-pi-subagent\docs\PRD\sub-agent-specs.xlsx"

HEADERS = ["ID", "Title", "What & Why", "How (API)", "Acceptance", "Failure modes", "Test plan", "LOC est"]

# ---------------------------------------------------------------------------
# Spec content (dict-based)
# ---------------------------------------------------------------------------

SPEC_01_COMMANDS = [
    dict(
        id="CMD-01",
        title="/loop-subagent slash command",
        what="A parallel slash command to /loop that creates a sub-agent loop. Mirrors /loop's `<interval> <prompt>` form so muscle memory transfers.",
        how="`/loop-subagent <interval> <prompt> [--goal <text>] [--success-criteria <text>] [--failure-criteria <text>] [--state-file <path>] [--model <name>] [--max-tokens <n>] [--max-iterations <n>] [--iteration-timeout <ms>]`. Backed by `src/commands/loop-subagent-command.ts`; the handler calls `LoopCreate({ isolation: 'sub-agent', ... })`.",
        accept="The command exists, accepts the same args as /loop plus the new flags, creates a LoopEntry with `isolation: 'sub-agent'`. Invoking with no args shows a TUI form. Tab completion lists the new flags. Invalid args produce a structured error referencing the flag.",
        fail="Invalid interval / cron expression; missing prompt; missing /loop-settings; user permission to create a loop.",
        test="test/loop-subagent-command.test.ts: parses each flag, persists to LoopStore, surfaces errors via textResult.",
        loc="150",
    ),
    dict(
        id="CMD-02",
        title="/loop-sub-agent-inspect slash command",
        what="Opens a TUI picker of recent iterations for a sub-agent loop, then a content viewer (result.md / session.jsonl / prompt.txt).",
        how="`/loop-sub-agent-inspect <loopId> [iterId]`. Implemented in `src/commands/loop-sub-agent-inspect-command.ts`. Falls back to the loop's latest 5 iterations when `iterId` is omitted. The viewer shells out to `$PAGER`, `code`, or `notepad` (Windows) in that order.",
        accept="Command exists; with no args shows a loop picker; with a loopId shows a recent-iterations picker; with a loopId+iterId shows the result picker (result / session / prompt / cost). The viewer opens the right file with the right tool. The picker navigates with arrow keys and exits on Esc / q.",
        fail="Loop not found; iteration not found; file missing; $PAGER not on PATH. Each surfaces a one-line error to the user.",
        test="test/loop-sub-agent-inspect.test.ts: argument parsing, picker state, file resolution.",
        loc="200",
    ),
    dict(
        id="CMD-03",
        title="/loop-sub-agent-stop slash command",
        what="Sends SIGTERM to one or all in-flight sub-agent iterations of a loop, with optional iterId to target one.",
        how="`/loop-sub-agent-stop <loopId> [iterId]`. Implemented in `src/commands/loop-sub-agent-stop-command.ts`. The handler calls `subAgentStore.kill({ loopId, iterId? })` which signals the child and finalizes the result with `status: 'cancelled'`.",
        accept="Command exists; with no args shows a loop picker. With loopId alone, kills all in-flight iterations of that loop. With loopId+iterId, kills that one. After kill, a `priority: 'urgent'` notification surfaces to the parent. The result.json is written with `status: 'cancelled'`.",
        fail="Loop not found; iteration already terminal; child pid no longer exists (already exited).",
        test="test/loop-sub-agent-stop.test.ts: kill signal, result.json finalization, notification emission.",
        loc="100",
    ),
    dict(
        id="CMD-04",
        title="/loop-cost slash command",
        what="Shows a table of per-loop and per-session cost & token totals for all sub-agent loops in this session.",
        how="`/loop-cost` (no args). Implemented in `src/commands/loop-cost-command.ts`. Reads from `subAgent/cost-tracker.ts` and prints a table. Output is plain text, not TUI, for easy copy-paste.",
        accept="Command exists; prints a table with columns Loop, Iterations, Last run, Total tokens, Total cost, plus a Total row. In-process loops show 0 cost (reserved column for future use).",
        fail="No sub-agent loops exist (print a 'no sub-agent activity yet' message).",
        test="test/loop-cost.test.ts: report shape, rounding, empty state.",
        loc="80",
    ),
    dict(
        id="CMD-05",
        title="LoopCreate / LoopUpdate / LoopDelete schema extensions",
        what="The three existing tools accept the new sub-agent fields. Backwards-compatible: existing callers pass no new fields and get the v2.x behaviour.",
        how="Extend `src/tools/loop-tools.ts`: add to `LoopCreate` the optional fields `isolation`, `goal`, `successCriteria`, `failureCriteria`, `stateFile`, plus a `subAgent` sub-object for per-loop overrides (model, maxTokens, maxIterations, iterationTimeoutMs, tools). `LoopUpdate` accepts the same fields, updates atomically. `LoopDelete` on a sub-agent loop sends SIGTERM to in-flight children and finalizes their results before the loop is removed.",
        accept="LoopCreate with no new fields works exactly as v2.x. LoopCreate with `isolation: 'sub-agent'` persists the new fields and shows them in LoopList. LoopUpdate with `goal: 'New goal'` updates only that field. LoopDelete on a sub-agent loop with 3 in-flight children finalizes all 3 with `status: 'cancelled'`.",
        fail="Invalid field values; type mismatches; concurrent updates race (existing LoopStore.withLock handles it).",
        test="test/loop-tools-extended.test.ts: each new field is persisted, restored, validated. Existing v2.x loop-tools tests pass unchanged.",
        loc="200",
    ),
    dict(
        id="CMD-06",
        title="LoopInspect tool",
        what="The agent-callable counterpart to /loop-sub-agent-inspect. Returns a structured summary so the agent can reason about its own sub-agent runs.",
        how="`subagent({ action: 'inspect', loopId, iterId? })` exposed as a new tool `LoopInspect({ loopId, iterId? })` registered in `src/tools/loop-tools.ts`. Returns `{ loop: LoopEntry, iteration: { id, status, startedAt, finishedAt, durationMs, tokens, costUsd, resultPath, childSessionPath, preview, errorMessage? } }`.",
        accept="The tool is registered, the agent can call it, the response is structured JSON. For a sub-agent loop, iteration is populated. For an in-process loop, iteration is the latest run summary (or null if none). Calling with a non-existent loopId returns a structured error `{ code: 'loop_not_found' }`.",
        fail="Loop not found; iteration not found; permission errors (none day 1).",
        test="test/loop-inspect-tool.test.ts: success path, missing loop, missing iter, response shape.",
        loc="150",
    ),
    dict(
        id="CMD-07",
        title="LoopList extended with sub-agent fields",
        what="LoopList output gains 3 new columns: `isolation`, `goal` (truncated), and `lastResult` (status + iterId + duration).",
        how="Extend the existing `LoopList` tool in `src/tools/loop-tools.ts` to include the new fields. Add a `filter: 'sub-agent' | 'in-process'` optional field so the user can narrow the list to sub-agent loops only.",
        accept="LoopList with no args shows every loop with the new columns. `LoopList({ filter: 'sub-agent' })` shows only sub-agent loops. The output is still a structured list, not a TUI render.",
        fail="None day 1.",
        test="test/loop-list-extended.test.ts: column population, filter behaviour, no breaking change to existing assertions.",
        loc="80",
    ),
]

SPEC_02_SCHEMA = [
    dict(
        id="SCH-01",
        title="isolation: 'in-process' | 'sub-agent' on LoopEntry",
        what="Switches the loop's execution mode. Default is 'in-process' (preserves v2.x behaviour).",
        how="Add to `src/types.ts`: `export type LoopIsolation = 'in-process' | 'sub-agent';`. Add to `LoopEntry`: `isolation?: LoopIsolation;` (optional for backwards compat; missing = 'in-process').",
        accept="A loop loaded from .pi/loops/loops.json with no `isolation` field behaves as `isolation: 'in-process'`. Setting it to 'sub-agent' triggers the new code path. Validation: must be one of the two values; reject unknown.",
        fail="Unknown value at write time; field is silently dropped if a future version adds a new value (we never delete old values).",
        test="test/loop-store-isolation.test.ts: round-trip with both values; missing field defaults correctly.",
        loc="40",
    ),
    dict(
        id="SCH-02",
        title="goal: string on LoopEntry",
        what="A free-text description of the loop's purpose. Surfaced in the FleetView, the inspect picker, and the cost report. Not evaluated by the runtime (it is documentation, not a behaviour driver).",
        how="Add to `LoopEntry`: `goal?: string;`. Validation: non-empty trimmed, max 1,000 chars, no NUL bytes.",
        accept="Persists correctly, round-trips through LoopStore, surfaces in the FleetView row and the cost report. Empty string is rejected; unset is allowed.",
        fail="Too long (reject with error); NUL bytes (reject); non-UTF-8 bytes (reject).",
        test="test/loop-schema-goal.test.ts: validation, round-trip, display.",
        loc="30",
    ),
    dict(
        id="SCH-03",
        title="successCriteria: string on LoopEntry",
        what="A free-text description of when an iteration counts as success. Evaluated by the parent after each sub-agent iteration against the child's `result.md`.",
        how="Add to `LoopEntry`: `successCriteria?: string;`. Validation: non-empty trimmed, max 2,000 chars. Storage: same as goal. Semantics: when the evaluator runs, it sees the successCriteria string; the result is either 'succeeded' (default) or 'succeeded-by-criteria' when the criteria match. The loop's overall status transitions to 'completed' when all sub-agent iterations satisfy the criteria AND `subAgent.maxIterations` is set; otherwise it remains 'active'.",
        accept="Persists correctly; round-trips through LoopStore; the evaluator consumes it; the loop auto-completes only when the criteria match AND maxIterations is set.",
        fail="Validation failures; criteria that match every iteration (loop completes too early); criteria that never match (loop never completes).",
        test="test/loop-schema-success.test.ts: validation, evaluator integration, completion flow.",
        loc="40",
    ),
    dict(
        id="SCH-04",
        title="failureCriteria: string on LoopEntry",
        what="A free-text description of when an iteration counts as a failure. Evaluated similarly to successCriteria; failures trigger backoff and a `priority: 'urgent'` notification.",
        how="Add to `LoopEntry`: `failureCriteria?: string;`. Validation: non-empty trimmed, max 2,000 chars. Semantics: when matched, the iteration is `status: 'failed_by_criteria'`, the loop's `consecutiveFailures` counter increments, and the scheduler applies backoff (default: skip the next cron tick). After 3 consecutive failures, the loop is marked `status: 'paused_failure'` and a `priority: 'urgent'` notification asks the user to intervene.",
        accept="Persists correctly; evaluator consumes it; backoff is applied after a match; 3 consecutive failures transition the loop to 'paused_failure' with an urgent notification.",
        fail="Validation failures; backoff too aggressive (the loop appears dead); 3-failure threshold is wrong for a noisy loop.",
        test="test/loop-schema-failure.test.ts: validation, backoff application, pause behaviour.",
        loc="50",
    ),
    dict(
        id="SCH-05",
        title="stateFile: string on LoopEntry",
        what="An optional path to a JSON file the child reads at the start of each iteration and writes at the end. Lets the loop maintain state across iterations without the child needing to know the loop exists.",
        how="Add to `LoopEntry`: `stateFile?: string;`. Validation: a path that resolves (relative to the loop's `loopScope` root), max 1,000 chars. Format: JSON object, max 256 KiB. The parent passes the path to the child as `--state-file <path>`. The runtime locks the file for concurrent access; if the child holds a stale lock for >5s, the parent takes the lock anyway and finalizes the iteration as `status: 'lock_timeout'`.",
        accept="Persists correctly; parent passes the path to the child; the child reads and writes it; the runtime locks concurrent access; stale lock is taken after 5s with `lock_timeout` finalization.",
        fail="Path doesn't exist (create it on first iteration); path is a directory; file size exceeds 256 KiB (reject); concurrent write from another process (lock + timeout).",
        test="test/loop-schema-state-file.test.ts: validation, locking, read/write round-trip via a mock child.",
        loc="80",
    ),
    dict(
        id="SCH-06",
        title="subAgent sub-object on LoopEntry",
        what="Per-loop overrides for the sub-agent runtime. Mirrors the global `subAgent` settings block but per-loop.",
        how="Add to `LoopEntry`: `subAgent?: { model?: string; thinking?: 'off' | 'low' | 'medium' | 'high'; tools?: string[]; iterationTimeoutMs?: number; iterationTokenBudget?: { in: number; out: number }; maxTokens?: number; maxIterations?: number; retainIterations?: number; label?: string; syncGit?: boolean; }`. Merge order: defaults from settings.subAgent < per-loop LoopEntry.subAgent < per-launch env overrides.",
        accept="Merges with defaults in the right order; all fields are optional; persists; round-trips. The runtime applies the merged config at spawn time.",
        fail="Missing required nested fields (allowed — they fall through to defaults); invalid model name (rejected at spawn time, not at write time); concurrent overrides.",
        test="test/loop-schema-subagent.test.ts: merge order, validation, defaults.",
        loc="80",
    ),
    dict(
        id="SCH-07",
        title="Backwards compatibility: existing v2.x loops",
        what="A loop loaded from .pi/loops/loops.json with no new fields is unchanged. All new fields are optional; missing means 'use the default'.",
        how="No change to LoopStore. The migration in `migration/v2-to-v2.5.ts` is a settings-only migration (see Spec 07), not a loop migration. Existing v2.x loop-tools tests pass unchanged.",
        accept="All existing v2.x loop-store, loop-tools, loop-reducer tests pass without modification. A loop with no new fields behaves identically to v2.x.",
        fail="None — this is a one-way compatibility guarantee.",
        test="All existing v2.x loop-store, loop-tools, loop-reducer tests pass without modification.",
        loc="0",
    ),
    dict(
        id="SCH-08",
        title="Validation rules & error messages",
        what="Every new field is validated at write time (LoopCreate, LoopUpdate, migration). The error message names the field and the constraint.",
        how="All validation lives in `src/runtime/loop-validation.ts` (a new file). Each field has a `validate*()` function returning either `{ ok: true, value }` or `{ ok: false, message }`. The store calls the validator before persisting. Migration also calls the validator on every record it touches.",
        accept="Every field has a validate*() function; bad data is rejected at write time with a message naming the field and the constraint; migration runs the same validator.",
        fail="Validator not called (silent bad data); validator accepts too much; validator rejects too much.",
        test="test/loop-validation.test.ts: every field's happy and sad paths; migration test that injects bad data and verifies it surfaces.",
        loc="120",
    ),
    dict(
        id="SCH-09",
        title="Atomic write through existing LoopStore",
        what="No new persistence. The new fields ride on the existing LoopEntry type and the existing write-tmp-then-rename pattern.",
        how="No change. The fields are added to the TypeScript type, validated, and persisted as part of the same JSON file. The existing `LoopStore.withLock()` provides the cross-process lock.",
        accept="Existing LoopStore write-tmp-then-rename covers this; no new persistence code; existing tests pass.",
        fail="None — uses the existing patterns.",
        test="No new test; existing LoopStore tests cover the write path.",
        loc="0",
    ),
]

SPEC_03_RUNTIME = [
    dict(
        id="RUN-01",
        title="spawnSubAgent() — child process spawner",
        what="Spawns a fresh child pi process with its own session file, prompt file, and tool allowlist. The child is the long-lived boundary: every iteration gets one child, runs to completion (or timeout), and exits.",
        how="`src/runtime/sub-agent/spawn.ts` exports `async function spawnSubAgent(req: SpawnRequest): Promise<SpawnHandle>`. `SpawnRequest` includes: loopId, iterId, cwd, childSessionPath, promptPath, model, thinking, tools, iterationTimeoutMs, piBinary, envOverrides. `SpawnHandle` exposes pid, childSessionPath, resultPath, startedAt, kill(), wait(). The implementation uses `child_process.spawn` with `detached: false` and `windowsHide: true`.",
        accept="A spawn creates the session file (empty, mode 0600 on Unix), invokes the pi binary with `--session-file <path> --prompt @<prompt.txt> --non-interactive --no-extensions --max-duration-ms <timeout>` plus the model / thinking / tools flags. The returned handle has the pid and a kill() method. A test that mocks `child_process.spawn` verifies the argv. A test that uses a real pi binary against a temp dir verifies the child runs and writes a session file.",
        fail="pi binary not on PATH; child crashes during init; spawn race (parent calls kill() before the child has started); ENV inheritance pollution.",
        test="test/sub-agent-spawn.test.ts: argv shape, env overrides, kill semantics, wall-clock timer (with vi.useFakeTimers).",
        loc="350",
    ),
    dict(
        id="RUN-02",
        title="subAgent-results/<loopId>/iter-<N>/ directory layout",
        what="One directory per iteration. Contains the child's session file, the prompt we sent, the result the parent wrote, and the child's stdout/stderr capture.",
        how="Layout: `<loopScope>/sub-agent-results/<loopId>/state.json` (per-loop ledger), `<loopScope>/sub-agent-results/<loopId>/iter-<N>/session.jsonl` (child session), `<loopScope>/sub-agent-results/<loopId>/iter-<N>/prompt.txt` (what we sent), `<loopScope>/sub-agent-results/<loopId>/iter-<N>/result.md` (child's own writeup), `<loopScope>/sub-agent-results/<loopId>/iter-<N>/result.json` (parent's view of the outcome), `<loopScope>/sub-agent-results/<loopId>/iter-<N>/child-stdout.log` and `child-stderr.log` (capped at 256 KiB each), `<loopScope>/sub-agent-results/<loopId>/iter-<N>/child.pid` (pid at spawn, deleted at finalize).",
        accept="The layout matches the spec above; every file is created in the right place; pid file is deleted at finalize; logs are capped at 256 KiB.",
        fail="Directory creation races (two iterations starting simultaneously); disk full; permissions.",
        test="test/sub-agent-dir-layout.test.ts: every expected file is created; pid file is deleted at finalize; logs are capped.",
        loc="100",
    ),
    dict(
        id="RUN-03",
        title="result.json schema",
        what="The parent's view of one iteration's outcome. Written by the parent on child exit (not by the child). The child writes `result.md`; the parent correlates that with the exit signal and writes result.json.",
        how="`src/runtime/sub-agent/result-store.ts` exports `finalize(loopId, iterId, childExit, childSessionTail): SubAgentResult`. `SubAgentResult` is the type from PRD §7.4: `schemaVersion, loopId, iterId, status, startedAt, finishedAt, durationMs, tokens, costUsd, exitCode, processSignal, resultPath, preview, errorMessage?, model?, thinking?, childSessionPath`. The write is atomic (tmp + rename).",
        accept="Schema is what PRD §7.4 defines; the parent is the only writer; the write is atomic; the file round-trips through the watcher and the cost-tracker.",
        fail="Result file is written by the child too (rejected — only the parent writes result.json); atomic write fails (parent re-tries 3 times then writes a tombstone); child session file is truncated (read whatever is there, mark status 'orphaned_partial').",
        test="test/sub-agent-result.test.ts: schema, atomic write, finalize race with concurrent kills.",
        loc="200",
    ),
    dict(
        id="RUN-04",
        title="prompt.txt: what we send to the child",
        what="The file the child reads as its initial user message. Written by the parent before spawn; survives crashes for inspection.",
        how="`src/runtime/sub-agent/prompt-writer.ts` writes `<iterationDir>/prompt.txt` with the loop's prompt wrapped in a small header that tells the child the loop name, iteration number, and any state-file path. The child gets the prompt as `@<prompt.txt>` via pi's @file syntax.",
        accept="The file exists at the right path with the right content (header + prompt); written before spawn; readable by the child; survives a parent crash.",
        fail="Prompt file write fails (parent aborts the spawn with a clear error); the prompt is empty (parent aborts); the child is told to write a result file but doesn't (finalize marks status 'failed_no_result').",
        test="test/sub-agent-prompt-writer.test.ts: header format, file existence, atomic write.",
        loc="100",
    ),
    dict(
        id="RUN-05",
        title="result-watcher.ts — observes child exit and finalizes",
        what="A long-running module that owns the in-flight iteration table. On child exit, reads the session file tail, calls costTracker.close(), resultStore.finalize(), and notificationRuntime.enqueue().",
        how="`src/runtime/sub-agent/result-watcher.ts` exposes `start(loopId)`, `stop(loopId)`, `spawn(loopId, iterId, handle)`, `kill(loopId, iterId?)`, `reconcile(sessionId, nowMs)`, `snapshot(sessionId)`. Internally: a `Map<loopId, Map<iterId, ActiveIteration>>` with a heartbeat per iteration. The heartbeat runs every 30s in `session-runtime.ts`'s existing 30s pump.",
        accept="A long-running module; start/stop/spawn/kill/reconcile/snapshot are all public; the in-flight table is correct; on child exit the watcher finalizes and enqueues a notification within 1s.",
        fail="Child is killed before the watcher can finalize (SIGKILL race); session file is locked by another process (read with retry + 1s timeout); finalize throws (the watcher's catch logs and surfaces a `priority: 'urgent'` notification).",
        test="test/sub-agent-result-watcher.test.ts: spawn/exit/finalize round-trip, kill-during-spawn, reconcile on simulated restart.",
        loc="450",
    ),
    dict(
        id="RUN-06",
        title="Parent-restart reconciliation",
        what="When the parent restarts, in-flight iterations may still be alive in the OS. The watcher walks the active iterations, checks each child's pid, and either reattaches (if alive) or finalizes as 'orphaned' (if dead).",
        how="On `session-runtime.ts` startup, call `resultWatcher.reconcile(sessionId, Date.now())`. For each `status: 'running'` iteration: read `lastHeartbeatAt`; if older than the stale threshold (default 5 min) and the pid is no longer alive, finalize as 'orphaned' (with the tokens consumed so far); if the pid is alive, attach a new exit handler to it (best-effort).",
        accept="On parent restart, all in-flight iterations are reconciled within 1 heartbeat; orphaned iterations are finalized with the right status; alive iterations get a new exit handler; the result.json is written exactly once.",
        fail="Pid reuse (the watcher uses a nonce stored in a sidecar file to verify the pid is the same child); orphan detection takes 5+ minutes (acceptable); child finished exactly as the parent restarted (race — both write to result.json; the parent uses atomic write + fsync to win the race).",
        test="test/sub-agent-restart-reconcile.test.ts: simulate parent restart with 3 in-flight iterations (1 alive, 1 dead, 1 finishing); assert each finalizes correctly.",
        loc="200",
    ),
    dict(
        id="RUN-07",
        title="Graceful shutdown handler",
        what="When the parent receives SIGINT or is being closed, the watcher cancels in-flight children and finalizes them as 'cancelled'.",
        how="On SIGINT, `resultWatcher.shutdown()` walks every active iteration, sends SIGTERM, waits up to 5s for graceful exit, then SIGKILL. The result.json is written with `status: 'cancelled'`. After shutdown, the parent proceeds to exit.",
        accept="On SIGINT, all in-flight children are SIGTERM'd, then SIGKILL'd if needed; the result.json is written with 'cancelled' status; double-shutdown is idempotent; the parent can exit cleanly.",
        fail="Child refuses to exit on SIGTERM (SIGKILL after 5s); shutdown called twice (idempotent); a child's last-gasp write arrives after the parent has already written result.json (parent uses 'last write wins' with a tombstone file to mark the cancellation as final).",
        test="test/sub-agent-shutdown.test.ts: SIGTERM-then-SIGKILL escalation, double-shutdown idempotency.",
        loc="120",
    ),
    dict(
        id="RUN-08",
        title="cost-tracker.ts — per-loop and per-session cost ledger",
        what="Tracks tokens and USD cost across all iterations of a loop and across all loops in a session. Source of truth for the cost report and the cost ceiling.",
        how="`src/runtime/sub-agent/cost-tracker.ts` exposes `open(loopId, iterId, model)`, `close(loopId, iterId, tokens, durationMs)`, `loopReport(loopId)`, `sessionReport()`, `cumulativeCostUsd()`. The per-iteration cost is computed from the model's published price table (a small JSON in the same file); the cumulative cost is the sum. The ledger is in-memory; the loop's `state.json` is the durable mirror.",
        accept="Tracks tokens + cost per iteration and cumulatively; reads/writes state.json; survives parent restart; session report aggregates all loops.",
        fail="Model not in the price table (fall back to a default price; warn the user); ledger diverges from on-disk state (parent re-reads state.json on startup); cost ceiling hit (the scheduler's gate catches it before the next spawn).",
        test="test/sub-agent-cost-tracker.test.ts: open/close round-trip, model price lookup, cumulative aggregation, divergence detection on simulated restart.",
        loc="200",
    ),
    dict(
        id="RUN-09",
        title="child-pid.ts — cross-platform owned-process tree",
        what="Tracks the pids of every child the parent has spawned, so the parent can kill them on shutdown or reconcile them on restart.",
        how="`src/runtime/sub-agent/child-pid.ts` exposes `track(loopId, iterId, pid, nonce)`, `untrack(loopId, iterId)`, `isAlive(pid)`, `all()`. `isAlive` uses `process.kill(pid, 0)` on Unix (signal 0 = existence check) and `Get-Process -Id <pid>` on Windows. The nonce is written to a sidecar file at spawn time; the reconciliation uses pid + nonce to detect pid reuse.",
        accept="Tracks pids + nonces; detects pid reuse via the nonce; isAlive works on both Unix and Windows; the tracking file is reload-safe.",
        fail="Pid reuse (the nonce catches it); Windows pid wraparound at 32 bits (the nonce catches it); the tracking file is corrupted on disk (the watcher skips and re-derives from the iteration dir contents).",
        test="test/sub-agent-child-pid.test.ts: tracking, isAlive, reconciliation nonce check, Windows path mocked via `child_process.exec`.",
        loc="150",
    ),
    dict(
        id="RUN-10",
        title="Cross-platform child process kill",
        what="SIGTERM on Unix maps to `TerminateProcess` on Windows, which is not a graceful signal. We use a two-stage kill (SIGTERM at T-30s, SIGKILL at T) to give the child a chance to flush.",
        how="Implemented inside `spawnSubAgent()`'s wall-clock timer. The timer is set at spawn time; on fire, it calls `child.kill('SIGTERM')`, waits 30s, and calls `child.kill('SIGKILL')` if the child is still alive. Node.js maps both signals to the appropriate OS primitive.",
        accept="Two-stage kill works on both Unix and Windows; SIGTERM at T-30s; SIGKILL at T; the result.json is written with the right status; cross-platform via Node's signal API.",
        fail="Windows: SIGTERM kills immediately with no flush opportunity (acceptable — the child writes to its session file as it goes, not in a single final write); the child handles SIGTERM and does not exit (SIGKILL after 30s); the parent crashes between SIGTERM and SIGKILL (the next parent restart's reconciliation marks the iteration 'orphaned').",
        test="test/sub-agent-kill.test.ts: SIGTERM-then-SIGKILL escalation, child-ignores-SIGTERM path, parent-crash-mid-kill recovery.",
        loc="100",
    ),
]

SPEC_04_SCHEDULER = [
    dict(
        id="SCH-01",
        title="Concurrency cap gate (activeIterationsMax)",
        what="Bounds the number of sub-agent iterations in flight across all sub-agent loops in this session.",
        how="`src/runtime/sub-agent/scheduler.ts` exports `gate(req: GateRequest): GateDecision`. The gate checks: (1) `activeCount >= settings.subAgent.activeIterationsMax` → defer; (2) loop's `consecutiveFailures >= 3` → pause; (3) `cumulativeCostUsd >= settings.subAgent.costCeilingUsd` → defer; (4) `cumulativeTokens >= loop.subAgent.maxTokens` → pause; (5) `iterCount >= loop.subAgent.maxIterations` → pause; (6) capability ceiling from pi-subagents (if installed) → policy_denied. Each failure returns a tagged `GateDecision` so the runtime can write the right notification.",
        accept="The cap is enforced; defer fires when the cap is reached; the deferred iteration is re-attempted on the next trigger tick or the next `agent_end`; the right notification is written.",
        fail="Race: two iterations complete simultaneously and free up slots; the gate sees the old count (acceptable — one extra spawn, within cap+1); the cap is set to 0 (interpret as 1, the floor).",
        test="test/sub-agent-scheduler-concurrency.test.ts: cap enforcement, race, defer notification.",
        loc="150",
    ),
    dict(
        id="SCH-02",
        title="Iteration cap gate (maxIterations)",
        what="Stops a loop after N successful or failed iterations.",
        how="The scheduler reads `loop.subAgent.maxIterations` and compares to `loop.iterCount` (a new field on LoopEntry, incremented in the watcher after finalize). When the cap is hit, the loop's status transitions to `paused_cap` and a `priority: 'urgent'` notification surfaces.",
        accept="Loop stops after N iterations; cap is honoured; status transitions to 'paused_cap' with an urgent notification.",
        fail="Cap hits mid-iteration (the in-flight iteration completes normally; the next one is the one that gets blocked); cap lowered at runtime (existing iterations still count toward the new cap).",
        test="test/sub-agent-scheduler-iter-cap.test.ts: cap hit, status transition, notification shape.",
        loc="80",
    ),
    dict(
        id="SCH-03",
        title="Token budget gate (maxTokens)",
        what="Stops a loop when its cumulative token usage (in + out, summed across all iterations) reaches a cap.",
        how="The scheduler reads `loop.subAgent.maxTokens` and compares to `loop.cumulativeTokens` (mirrored from costTracker). The check is at spawn time, not at runtime (the iteration in progress is allowed to finish; the next spawn is blocked). When the cap is hit, status becomes `paused_budget`.",
        accept="Loop stops at budget; cap is honoured; the in-flight iteration is allowed to finish; status transitions to 'paused_budget'.",
        fail="Tokens reported late (the spawn proceeds, but the gate catches it before commit); user raises the budget (the new cap is honoured on the next spawn).",
        test="test/sub-agent-scheduler-budget.test.ts: cap hit, mid-iteration behaviour, raise-budget round-trip.",
        loc="80",
    ),
    dict(
        id="SCH-04",
        title="Defer / pause / fail behaviours",
        what="When a gate fails, the runtime must decide between three behaviours: defer (retry on next trigger or next agent_end), pause (stop firing until user intervention), or fail (mark the loop as `failed` and stop).",
        how="Defined in the scheduler: cap-hit → pause; budget-hit → pause; concurrency-cap → defer; policy-denied → defer with a `priority: 'urgent'` notification; cap reached naturally → pause with `priority: 'urgent'`. The behaviours are spelled out in a small enum and the scheduler returns the right one. The watcher applies it (sets status, writes notification).",
        accept="Each gate-condition maps to the right behaviour; the right notification priority is used; the status transition is durable.",
        fail="Wrong behaviour for a gate condition (defer when it should pause); notification priority is wrong; status transition is not durable.",
        test="test/sub-agent-scheduler-behaviours.test.ts: each gate-condition × each expected behaviour; notification priority assertion.",
        loc="100",
    ),
    dict(
        id="SCH-05",
        title="Result evaluator — success / failure matching",
        what="Applies the loop's `successCriteria` and `failureCriteria` to the child's `result.md` after each iteration. Returns 'matched-success' / 'matched-failure' / 'no-match'.",
        how="`src/runtime/sub-agent/evaluator.ts` exposes `evaluate(resultMd: string, successCriteria?: string, failureCriteria?: string): { success: boolean; failure: boolean; reason?: string }`. The default evaluator is regex-based: if `successCriteria` is set, treat it as a regex and match against `result.md`; same for failure. If a criteria is unset, that side defaults to 'no-match' (so the iteration is treated as succeeded by default unless the failure criteria match).",
        accept="Applies criteria to result.md; returns matched-success / matched-failure / no-match; handles missing criteria; catches invalid regex.",
        fail="Criteria is invalid regex (catch the error, treat as 'no-match', log a warning); result.md is empty (treat as 'no-match', no error); criteria are contradictory (the failure check runs second; both can match, the failure wins).",
        test="test/sub-agent-evaluator.test.ts: regex happy paths, invalid regex, empty result, contradictory criteria.",
        loc="150",
    ),
    dict(
        id="SCH-06",
        title="Optional LLM-call evaluator (stretch)",
        what="A v2.5.1 stretch: an evaluator that calls the parent's LLM with the criteria + result and asks for a verdict. More expensive (extra LLM call per iteration) but more accurate than regex.",
        how="If `successCriteria` or `failureCriteria` is set AND the user has enabled `subAgent.useLlmEvaluator: true` in settings, route the criteria through a small LLM call after the regex evaluator. The LLM verdict wins when the two disagree. The call uses a fast/cheap model (haiku by default) and is capped at 1k tokens.",
        accept="LLM verdict path; 5s timeout; on timeout, fall back to the regex verdict; on disagreement, the LLM verdict wins; the model is configurable.",
        fail="The LLM call is too slow (cap at 5s; if it overruns, fall back to the regex verdict); the LLM hallucinates a match (the user sees the result.md and the verdict in the wake, can intervene).",
        test="test/sub-agent-evaluator-llm.test.ts: LLM-verdict path, timeout fallback, disagreement handling. (Mocked LLM call; no real LLM in tests.)",
        loc="180",
    ),
    dict(
        id="SCH-07",
        title="stateFile read / write semantics",
        what="The child reads the state file at the start of each iteration and writes it at the end. The parent passes the path via `--state-file`. The runtime locks the file for concurrent access; on stale lock, the parent takes the lock and finalizes as 'lock_timeout'.",
        how="The child gets a `stateFile` tool (added to its allowlist if not present; the spawn adds it). The tool reads the JSON, the child can mutate it, and writes it back atomically. The runtime uses a separate `stateFile.lock` sidecar; if the lock is held >5s, the parent takes it. The state file format is a JSON object, max 256 KiB.",
        accept="Read/write round-trip; locking works; malformed recovery; size cap enforced; lock_timeout after 5s on stale lock.",
        fail="State file is malformed (the child gets an empty state and a warning; the parent finalizes with status 'succeeded' but flags the malformed state in result.json); child forgets to write the state file (parent reads whatever is there, accepts the previous state); state file grows unboundedly (caller's responsibility — the runtime enforces 256 KiB max).",
        test="test/sub-agent-state-file.test.ts: read/write round-trip, locking, malformed recovery, size cap.",
        loc="180",
    ),
    dict(
        id="SCH-08",
        title="Backoff after failure",
        what="After a failed iteration, the loop's next cron tick is skipped (or delayed). Backoff config: default is 'skip next tick'. User can set `subAgent.backoff: { strategy: 'fixed' | 'exponential', baseMs?: number, maxMs?: number }`.",
        how="The watcher, on finalizing a failure, sets `loop.nextEligibleAt = max(now, lastFiredAt + backoffMs)`. The trigger system checks `nextEligibleAt` before firing and skips if it hasn't been reached. Exponential backoff: 1m, 5m, 25m, 1h, 1h cap.",
        accept="Skip-next-tick semantics work; exponential growth (1m, 5m, 25m, 1h); reset on success; backoff cleared via LoopUpdate.",
        fail="Backoff makes the loop appear dead (the user is told via the wake 'Loop #4 backoff: next fire at 14:55, last failure at 14:50'); backoff cleared on a successful iteration; backoff config is invalid (fall back to default).",
        test="test/sub-agent-backoff.test.ts: skip-next-tick semantics, exponential growth, reset on success, clearing via LoopUpdate.",
        loc="100",
    ),
    dict(
        id="SCH-09",
        title="Goal-driven loop completion (successCriteria + maxIterations)",
        what="If a loop has BOTH `successCriteria` and `subAgent.maxIterations`, it completes (status: 'completed') when either: an iteration satisfies the success criteria, OR the iteration cap is reached. This is the only way a sub-agent loop auto-completes.",
        how="The watcher, on finalize, checks: if success matched OR iterCount >= maxIterations, set loop.status = 'completed', write a `priority: 'urgent'` notification with the completion reason. Without both fields set, the loop never auto-completes — only the user can complete it via LoopDelete.",
        accept="Auto-completes when criteria match OR cap reached; writes the completion notification; without both fields set, the loop never auto-completes.",
        fail="Success criteria is too loose (the loop completes on the first iteration); maxIterations is too high (the loop runs for a long time before the cap is checked); neither is set (the loop never auto-completes, by design).",
        test="test/sub-agent-completion.test.ts: success-path completion, cap-path completion, no-fields-set persistence.",
        loc="120",
    ),
]

SPEC_05_TUI = [
    dict(
        id="TUI-01",
        title="formatSubAgentResult() — one-line summary formatter",
        what="Builds the wake string the parent sees. Tiered by priority: 200 chars for normal/urgent/defer, 1,000 chars for critical.",
        how="`src/runtime/sub-agent/notification-formatter.ts` exports `formatSubAgentResult(n: SubAgentResultNotification, settings): string`. The formatter is pure (no I/O). The output is plain text with backticks around paths. Critical previews are allowed to be longer; non-critical are terse.",
        accept="Format produces a one-line summary for normal, urgent, defer; up to 1,000 chars for critical. Paths are relative to the cwd. Tokens are formatted with thousands separators. The result is deterministic given the same input.",
        fail="Missing fields (preview, tokens, etc.) fall back to 'unknown'; result.md is too long (truncate at 200 / 1,000 chars with an ellipsis); iteration is in a non-terminal state (the formatter is only called on finalize, so this should not happen, but the formatter defends against it).",
        test="test/sub-agent-notification-formatter.test.ts: each priority tier, each failure status, truncation behaviour.",
        loc="120",
    ),
    dict(
        id="TUI-02",
        title="Notification queue integration (new kind: sub-agent-result)",
        what="The existing priority queue grows a new notification kind. No new wake mechanism — sub-agent results enter the same queue and respect the same flush paths.",
        how="Extend `ReducerNotification` in `src/types.ts` to include `kind: 'loop-fire' | 'sub-agent-result' | 'monitor-event' | ...`. Extend the notification formatter to dispatch on kind. Extend the existing priority-flush and idle-flush paths to handle the new kind. ADR-005's defer / normal / urgent / critical semantics apply unchanged.",
        accept="New kind enters the existing queue; respects priority flush; respects idle flush; coalesces by (loopId, iterId); each item is capped at 10 KiB.",
        fail="Two kinds in the same queue (the formatter dispatches); critical previews inflate the queue (capped at 10 KiB per item, regardless of preview length); the queue is full (the new item is coalesced by `(loopId, iterId)` — duplicates drop).",
        test="test/sub-agent-notification-queue.test.ts: enqueue, priority flush, idle flush, coalescing by key, queue cap.",
        loc="100",
    ),
    dict(
        id="TUI-03",
        title="Critical-priority interrupt gating",
        what="A critical-priority sub-agent result can interrupt the parent, but only if the parent is not currently processing another loop wake (avoid wake-interrupts-wake).",
        how="In `session-runtime.ts`'s existing `REQUEST_URGENT_FLUSH` pump: before delivering, check if the current turn is a loop wake. If yes, drop the interrupt and let the next idle flush handle it. If no, deliver. The check is a single boolean read of `agentRunning` + a tag on the running tool call (set by the wake handler).",
        accept="Critical sub-agent result can interrupt the parent unless the parent is in a loop wake; the user can opt into pure 'always interrupt' via `subAgent.criticalInterruptsAll: true`; the tag is read correctly.",
        fail="The tag is wrong (we deliver a critical wake while a normal wake is running — acceptable since critical is rarer); the user wants pure 'always interrupt' (set `subAgent.criticalInterruptsAll: true` to opt in); the parent is idle and the critical wake arrives (delivered immediately, no change from today).",
        test="test/sub-agent-critical-interrupt.test.ts: interrupt path, no-interrupt path, opt-in via setting.",
        loc="80",
    ),
    dict(
        id="TUI-04",
        title="Sub-agent FleetView belowEditor panel",
        what="A second widget registered below the editor, mirroring the existing LoopWidget. Shows live progress for all active sub-agent iterations. Hidden by default; appears when at least one iteration is in 'running' or 'starting' state.",
        how="`src/ui/sub-agent-fleet.ts` exports the widget. Registration in `src/index.ts` follows the existing pattern (`ctx.ui.setWidget('loops-sub-agent', factory, { placement: 'belowEditor' })`). The widget is hidden when `resultStore.activeCount() === 0`. Each row: `loop #N iter-M role duration tokens status`.",
        accept="Widget appears when iterations are running; hidden otherwise; renders throttle at 1 Hz; truncates to 5 visible rows with a '...N more' footer when there are >5 active.",
        fail="The widget flickers (rendering is throttled to 1 Hz); many active iterations (the panel truncates to 5 visible rows with a '...N more' footer); the user has 25 active iterations (truncation shows); the widget is hidden but should show (a small race in the active-count check; the watcher flushes on every finalize).",
        test="test/sub-agent-fleet-widget.test.ts: render output, hidden/shown transitions, throttle.",
        loc="250",
    ),
    dict(
        id="TUI-05",
        title="Ctrl+Shift+S FleetView overlay + Ctrl+Shift+C cost report",
        what="Two new keybindings. Ctrl+Shift+S opens a scrollable list of all known sub-agent iterations (active + last 50 terminal). Ctrl+Shift+C opens the /loop-cost report in a scrollable modal.",
        how="`src/ui/sub-agent-overlays.ts` exports `showSubAgentFleetOverlay()` and `showCostReportOverlay()`. Keybindings registered in `src/runtime/session-runtime.ts`'s existing keybinding setup. ADR-004's overlay-key naming convention applies; the key 'f' is unused inside the overlay (in-FleetView: f = filter, s = stop, i = inspect).",
        accept="Keybindings open the right overlay; each overlay is scrollable; the FleetView overlay filters; the cost report shows the same data as /loop-cost; Esc closes.",
        fail="The keybinding conflicts with another extension (the user is told at startup, per the existing pattern in pi-subagents); the overlay is open when an iteration finalizes (the overlay refreshes via the existing render loop).",
        test="test/sub-agent-overlays.test.ts: open/close, navigation, filter.",
        loc="300",
    ),
    dict(
        id="TUI-06",
        title="Status line cost indicator + slash command pickups",
        what="The editor's status line shows the session's cumulative sub-agent cost. Slash command /loop-sub-agent-stop and /loop-sub-agent-inspect are reachable from the panel via single-key shortcuts.",
        how="The status line update is one of two: append `sub-agent: $X.XX (N iter)` to the existing model / token line. Implemented as a hook in `src/ui/widget-render.ts` that reads `costTracker.sessionReport()` on every heartbeat. The single-key shortcuts are wired in the panel's `keypress` handler.",
        accept="Status line shows cumulative cost; truncated at 40 chars; hidden via `subAgent.showCostInStatusLine: false`; single-key shortcuts work from the panel.",
        fail="The status line is too long (truncate the sub-agent part with an ellipsis when it exceeds 40 chars); the user wants the cost hidden (add `subAgent.showCostInStatusLine: false` in settings).",
        test="test/sub-agent-statusline.test.ts: format, truncation, hidden setting.",
        loc="80",
    ),
]

SPEC_06_BRIDGE = [
    dict(
        id="BRG-01",
        title="registerBackgroundWorkProvider shim",
        what="If pi-subagents is installed, register pi-loop's active sub-agent iterations as background-work items so they show in /subagents-fleet and subagent_wait.",
        how="`src/runtime/sub-agent/pi-subagents-bridge.ts` calls `import('pi-subagents/background-work')` lazily. On success, registers a `BackgroundWorkProvider` with `name: 'pi-loop-sub-agent'`, `wakeChannels: ['loop:sub-agent:result']`, and a `listActiveWork()` callback that reads from `subAgentStore.snapshotActive(sessionId)`. The `reconcile()` callback is a no-op (we don't need a hook for in-process reconciliation).",
        accept="When pi-subagents is installed and a sub-agent iteration is active, the iteration shows up in /subagents-fleet. When the iteration finishes, it disappears. The registration is reload-safe (a re-spawn of the same provider replaces the old one).",
        fail="pi-subagents is not installed (no-op, no error); pi-subagents is installed but the import fails (silent no-op, no crash); the provider name collides with another extension (the registration fails; we surface a warning to the parent at startup, not an error).",
        test="test/sub-agent-bridge-register.test.ts: import success, listActiveWork shape, reconcile, registration failure surfaces a warning.",
        loc="200",
    ),
    dict(
        id="BRG-02",
        title="Capability ceiling auto-read",
        what="If pi-subagents has registered a capability ceiling via `registerSubagentCapabilityCeiling`, pi-loop intersects it with each loop's tool allowlist at spawn time. A first-time warning surfaces the intersection so the user knows it's happening.",
        how="`src/runtime/sub-agent/scheduler.ts` reads `getActiveSubagentCapabilityCeiling(sessionId)` from pi-subagents (lazy import). If a ceiling is active, intersect `loop.subAgent.tools` (or the default) with `ceiling.allowedTools`. If the intersection is empty, the spawn is deferred with reason 'policy_denied'. A warning is emitted the first time per session per loop via `pi.events.emit('loop:sub-agent:policy-warning', {...})`.",
        accept="Reads the ceiling; intersects with the allowlist; emits a warning the first time; defers on empty intersection; respects the opt-out setting.",
        fail="The ceiling changes mid-iteration (the in-flight iteration finishes with the old allowlist; the next spawn uses the new one — acceptable); the ceiling is overly restrictive (the user sees the policy_denied defer and can act); the user doesn't want this behaviour (set `subAgent.honorCapabilityCeiling: false` in settings).",
        test="test/sub-agent-bridge-ceiling.test.ts: intersection, defer on empty, warning emission, opt-out.",
        loc="120",
    ),
    dict(
        id="BRG-03",
        title="Feature detection via Symbol.for",
        what="The bridge only activates when `Symbol.for('pi-subagents.background-work.v1')` is registered (i.e. pi-subagents has been loaded). On first import attempt, set the symbol locally if it isn't set; if it is set, read from it.",
        how="The bridge module is loaded lazily. The first time a sub-agent spawn happens, it tries `await import('pi-subagents/background-work')`. If the import resolves, the symbol check + register call happens. If the import throws (module not found), the bridge is a no-op for the rest of the session.",
        accept="Feature-detected via Symbol.for; no-op when absent; no-op when import fails; safe against a false-positive package of the same name.",
        fail="False positive: a package named 'pi-subagents/background-work' exists but is not the same one (the symbol check catches it; the import is allowed to fail later, but the first import succeeds — we use the symbol check as the authority, not the import result).",
        test="test/sub-agent-bridge-feature-detect.test.ts: symbol absent, symbol present, false positive.",
        loc="60",
    ),
    dict(
        id="BRG-04",
        title="Wake channel integration",
        what="The bridge exposes 'loop:sub-agent:result' as a wake channel so pi-subagents can wake the parent faster than its 30s poll.",
        how="In the bridge registration, declare `wakeChannels: ['loop:sub-agent:result']`. The watcher emits this event on every finalize. pi-subagents' wait-completions subscribes and uses it to short-poll.",
        accept="Wake channel declared; emit on finalize; pi-subagents' wait-completions receives the event and short-polls.",
        fail="Event not delivered (the watcher falls back to the 30s poll); the event is delivered but pi-subagents isn't listening (no-op).",
        test="test/sub-agent-bridge-wake.test.ts: emit on finalize, channel declaration.",
        loc="30",
    ),
    dict(
        id="BRG-05",
        title="Lifecycle integration (reload, /resume)",
        what="After a /reload or /resume, the bridge re-registers itself with the current session's state. The previously active iterations are restored as background-work items.",
        how="On every session-runtime startup, the bridge re-registers. The `listActiveWork()` callback is called by pi-subagents' reconciler; the bridge returns the current snapshot. The reconciler merges with its own state.",
        accept="Re-registers on /reload and /resume; the previously active iterations are visible; reload-safe (the new provider replaces the old); no-op when pi-subagents isn't loaded.",
        fail="Re-registration replaces the old provider (reload-safe, per the pi-subagents contract); the new session has no active iterations (empty list returned, no error); pi-subagents is not loaded in the new session (no-op).",
        test="test/sub-agent-bridge-lifecycle.test.ts: reload, resume, no-active state.",
        loc="50",
    ),
    dict(
        id="BRG-06",
        title="Opt-out setting",
        what="A user who doesn't want the bridge can disable it via `subAgent.registerBackgroundWorkProvider: false` in settings.",
        how="The bridge checks the setting before registering. The default is true (auto-register on first spawn).",
        accept="Setting respected; default applied; no warning when opted out; no error when opted out.",
        fail="Setting is set to false (no-op, no warning); setting is missing (default true); setting is true but the import fails (no-op, no warning).",
        test="test/sub-agent-bridge-optout.test.ts: setting respected, default applied.",
        loc="20",
    ),
]

SPEC_07_SETTINGS_TESTS = [
    dict(
        id="SET-01",
        title="subAgent block in .pi/pi-loop-settings.json",
        what="The unified settings file (v2.0) gains a top-level `subAgent` key. All fields optional with safe defaults.",
        how="Add to `src/settings.ts`: `interface SubAgentSettings { defaultIsolation?: 'in-process' | 'sub-agent'; activeIterationsMax?: number; defaultIterationTimeoutMs?: number; defaultIterationTokenBudget?: { in: number; out: number }; piBinary?: string; envOverrides?: Record<string, string>; registerBackgroundWorkProvider?: boolean; honorCapabilityCeiling?: boolean; useLlmEvaluator?: boolean; costCeilingUsd?: number; showCostInStatusLine?: boolean; criticalInterruptsAll?: boolean; }`. The runtime merges this with built-in defaults (built-in first, user override).",
        accept="A fresh install with no `subAgent` block in settings.json behaves with all built-in defaults. A user can override any field; unknown fields cause a startup error (existing settings validation).",
        fail="Unknown field (rejected at startup); partial override (the rest is filled by defaults); concurrent settings writes (existing withLock handles it).",
        test="test/settings-subagent.test.ts: defaults, partial override, unknown-field rejection, merge order.",
        loc="100",
    ),
    dict(
        id="SET-02",
        title="v2-to-v2.5 one-shot migration",
        what="On first startup with v2.5, the migration adds the `subAgent` block to .pi/pi-loop-settings.json with all defaults. Loops are not migrated (existing loops are unaffected by the new fields).",
        how="`src/migration/v2-to-v2.5.ts` exports `run(projectRoot, settingsPath): { changed: boolean }`. The migration is idempotent: running it twice is a no-op. It only writes the file if the block is missing; it does not touch any other field.",
        accept="Migration is idempotent; only adds the subAgent block; doesn't touch other fields; errors are surfaced; the migration is no-op on a clean install.",
        fail="Settings file is read-only (the migration aborts with a clear error); the block exists but is partial (no-op — the existing block is honoured); the migration runs concurrently (existing withLock).",
        test="test/migration-v2-to-v2.5.test.ts: idempotency, no-other-fields-touched, error paths.",
        loc="80",
    ),
    dict(
        id="SET-03",
        title="/loop-settings TUI extensions",
        what="The /loop-settings TUI gains a 'Sub-agent defaults' sub-menu with: defaultIsolation, activeIterationsMax, defaultIterationTimeoutMs, piBinary. The other subAgent fields are advanced and require direct JSON editing.",
        how="Extend `src/commands/settings-command.ts`. The new sub-menu is the 5th item in the top-level menu. Each field is a cyclic form (same pattern as existing settings).",
        accept="The sub-menu exists; each field is editable; the saved settings round-trip through the file. The cyclic forms respect the existing format. Direct JSON editing still works for the advanced fields.",
        fail="The sub-menu is hidden when no sub-agent loops exist (no — always show, per the existing pattern of 'show settings that exist'); the user cancels mid-edit (existing cancellation path).",
        test="test/settings-command-subagent.test.ts: sub-menu render, each field's cyclic form, persistence.",
        loc="150",
    ),
    dict(
        id="TST-01",
        title="Unit test inventory",
        what="Every spec has at least one unit test. The full inventory is enumerated in the test plan appendix.",
        how="Tests co-located in test/ (per existing convention). Naming: `<module>.test.ts`. Coverage target: 80% line coverage for the new sub-agent/ directory.",
        accept="All tests pass; coverage report uploaded; no flakies over 100 runs.",
        fail="Coverage falls below 80% (CI gate); flaky tests; tests that mock too much (no real signal).",
        test="All tests pass; coverage report uploaded; no flakies over 100 runs.",
        loc="covered per spec",
    ),
    dict(
        id="TST-02",
        title="Integration test inventory",
        what="End-to-end tests that span spawn → run → result → notification. The integration test harness uses a real `pi` binary against a temp cwd.",
        how="test/integration/sub-agent-spawn.test.ts; test/integration/sub-agent-restart.test.ts; test/integration/sub-agent-concurrency.test.ts; test/integration/sub-agent-pi-subagents-bridge.test.ts. Each test creates a temp cwd, writes a fake loop prompt, calls the runtime, and asserts on the file outputs and the notification queue.",
        accept="All integration tests pass on Windows, macOS, Linux. CI runs them on every PR.",
        fail="A real `pi` binary is not on PATH (the test harness detects this and skips with a clear message); the test runs longer than the CI timeout (each test is capped at 30s); the harness leaks state between tests (each test uses a fresh tempdir).",
        test="All integration tests pass on Windows, macOS, Linux. CI runs them on every PR.",
        loc="covered per spec",
    ),
    dict(
        id="TST-03",
        title="E2E test inventory",
        what="TUI-driven tests that open a child terminal, drive /loop-subagent, and assert on the panel and overlay behaviour.",
        how="test/e2e/loop-subagent-tui.test.ts uses a child terminal driver (existing in the repo's test harness). The test: open the TUI, run /loop-subagent, watch the panel populate, press Ctrl+Shift+S, navigate, press 'i' to inspect, assert the result file opens.",
        accept="All E2E tests pass.",
        fail="TUI driver is flaky (skip on known-flaky hosts); the test is too slow (capped at 60s).",
        test="All E2E tests pass.",
        loc="covered per spec",
    ),
    dict(
        id="TST-04",
        title="Cross-platform verification",
        what="Every Windows-specific path is exercised on Windows. The full test suite runs on macOS, Linux, and Windows in CI.",
        how="CI matrix: ubuntu-latest, macos-latest, windows-latest. Each runs `npm run lint && npm run typecheck && npm test && npm run build`. The integration tests are gated on a `pi` binary being on PATH; the CI runners install it via a small `bin/install-pi.sh` script.",
        accept="All platforms green in CI for 7 consecutive days before v2.5.0 ships.",
        fail="A platform-specific bug (the test passes on macOS, fails on Windows); CI runtime exceeds the budget (the integration tests are the slow ones; parallelize across 4 shards).",
        test="All platforms green in CI for 7 consecutive days before v2.5.0 ships.",
        loc="covered per spec",
    ),
    dict(
        id="TST-05",
        title="Manual test scenarios",
        what="User-flow documents in userflow/ describe the manual scenarios. The user runs each before signing off on a release.",
        how="userflow/loop-subagent-create.md, loop-subagent-inspect.md, loop-subagent-stop.md, loop-subagent-cost.md, loop-subagent-restart-recovery.md. Each is a step-by-step script with expected output at each step.",
        accept="All 5 user-flows verified by the user before v2.5.0 ships.",
        fail="A user-flow step is wrong (update the flow doc); the user's environment differs (the flow doc names the environment).",
        test="All 5 user-flows verified by the user before v2.5.0 ships.",
        loc="covered per spec",
    ),
]

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
SHEETS = [
    ("01 Commands & tools", SPEC_01_COMMANDS),
    ("02 LoopEntry schema", SPEC_02_SCHEMA),
    ("03 Execution runtime", SPEC_03_RUNTIME),
    ("04 Scheduler & evaluator", SPEC_04_SCHEDULER),
    ("05 Notifications & TUI", SPEC_05_TUI),
    ("06 pi-subagents bridge", SPEC_06_BRIDGE),
    ("07 Settings migration tests", SPEC_07_SETTINGS_TESTS),
]

# Styles
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Calibri", size=10, color="000000")
ID_FONT = Font(name="Calibri", size=10, color="000000", bold=True)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOC_FILL = PatternFill("solid", fgColor="E7E6E6")

wb = Workbook()
wb.remove(wb.active)

# Build the cover sheet first so it ends up at index 0
cover = wb.create_sheet("Cover")
cover["A1"] = "Sub-agent feature specifications"
cover["A1"].font = Font(name="Calibri", size=18, bold=True, color="305496")
cover["A2"] = "Companion to docs/PRD/sub-agent.md (12k-word high-level PRD) and docs/PRD/sub-agent-questions.xlsx / -r2.xlsx (the decision matrices)"
cover["A2"].font = Font(name="Calibri", size=11, italic=True, color="404040")
cover["A2"].alignment = WRAP
cover["A4"] = "How to use this workbook"
cover["A4"].font = Font(name="Calibri", size=12, bold=True, color="305496")
cover["A5"] = (
    "Seven spec sheets, one per feature group. Each row is a self-contained requirement. The columns are the same on every sheet:\n\n"
    "  A  ID            short identifier (CMD-01, SCH-02, ...)\n"
    "  B  Title         short, 4-8 words\n"
    "  C  What & Why    the user-visible purpose (1-3 sentences)\n"
    "  D  How (API)     the function / command / setting / type signature\n"
    "  E  Acceptance    testable criteria (the spec's 'definition of done')\n"
    "  F  Failure modes what can go wrong and how we handle it\n"
    "  G  Test plan     which test file + key assertions\n"
    "  H  LOC est       rough new code size (a planning aid, not a contract)"
)
cover["A5"].alignment = WRAP
cover.row_dimensions[5].height = 200
cover.column_dimensions["A"].width = 110

# Build each spec sheet
for sheet_name, rows in SHEETS:
    ws = wb.create_sheet(sheet_name)
    ws.append(HEADERS)
    for c, _ in enumerate(HEADERS, 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for d in rows:
        # Build the 8-tuple from the dict; this is the point — no field can be silently dropped.
        rid = d["id"]
        title = d["title"]
        what = d["what"]
        how = d["how"]
        accept = d["accept"]
        fail = d["fail"]
        test = d["test"]
        loc = d["loc"]
        # Sanity: every key is present
        for k in ("id", "title", "what", "how", "accept", "fail", "test", "loc"):
            assert k in d, f"{sheet_name} {rid}: missing key {k}"
        ws.append([rid, title, what, how, accept, fail, test, loc])
        r = ws.max_row
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
        ws.cell(r, 1).font = ID_FONT
        ws.cell(r, 8).fill = LOC_FILL
        ws.cell(r, 8).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        max_len = max(len(str(ws.cell(r, c).value or "")) for c in range(2, 8))
        ws.row_dimensions[r].height = max(60, min(300, max_len // 4))

    widths = {1: 8, 2: 28, 3: 50, 4: 55, 5: 50, 6: 45, 7: 35, 8: 9}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "70AD47"

# Now fill in the cover's sheet index
cover["A12"] = "Sheet index"
cover["A12"].font = Font(name="Calibri", size=12, bold=True, color="305496")

cover["A14"] = "Sheet"
cover["A14"].font = HEADER_FONT
cover["A14"].fill = HEADER_FILL
cover["A14"].alignment = HEADER_ALIGN
cover["B14"] = "Purpose"
cover["B14"].font = HEADER_FONT
cover["B14"].fill = HEADER_FILL
cover["B14"].alignment = HEADER_ALIGN
cover["C14"] = "Requirements"
cover["C14"].font = HEADER_FONT
cover["C14"].fill = HEADER_FILL
cover["C14"].alignment = HEADER_ALIGN

INDEX_ROWS = [
    ("01 Commands & tools", "/loop-subagent slash command, /loop-sub-agent-inspect, /loop-sub-agent-stop, /loop-cost, LoopCreate/Update/Delete schema, LoopInspect tool, LoopList extended", 7),
    ("02 LoopEntry schema", "isolation, goal, successCriteria, failureCriteria, stateFile, subAgent sub-object, validation, backwards compat, atomic write", 9),
    ("03 Execution runtime", "spawnSubAgent, dir layout, result.json, prompt.txt, result-watcher, parent-restart reconcile, graceful shutdown, cost-tracker, child-pid, cross-platform kill", 10),
    ("04 Scheduler & evaluator", "concurrency cap, iteration cap, token budget, defer/pause/fail, regex evaluator, LLM evaluator (stretch), stateFile, backoff, completion", 9),
    ("05 Notifications & TUI", "formatSubAgentResult, notification queue integration, critical interrupt, FleetView panel, overlays, status line", 6),
    ("06 pi-subagents bridge", "background-work provider, capability ceiling, feature detection, wake channels, lifecycle, opt-out", 6),
    ("07 Settings migration tests", "settings schema, v2-to-v2.5 migration, /loop-settings TUI, unit/integration/E2E/cross-platform/manual test inventory", 8),
]
r = 15
for name, purpose, count in INDEX_ROWS:
    cover.cell(r, 1, name).font = Font(name="Calibri", size=10, bold=True)
    cover.cell(r, 1).alignment = WRAP
    cover.cell(r, 2, purpose).alignment = WRAP
    cover.cell(r, 2).font = BODY_FONT
    cover.cell(r, 3, str(count)).alignment = Alignment(horizontal="center", vertical="top")
    cover.cell(r, 3).font = BODY_FONT
    cover.row_dimensions[r].height = 60
    for c in (1, 2, 3):
        cover.cell(r, c).border = BORDER
    r += 1

cover.cell(r, 1, "TOTAL").font = Font(name="Calibri", size=10, bold=True)
cover.cell(r, 1).fill = LOC_FILL
cover.cell(r, 2, "55 requirements across 7 sheets. Total rough new code: ~4,300 LOC + 1,900 LOC of tests + 500 LOC of docs.").alignment = WRAP
cover.cell(r, 2).font = Font(name="Calibri", size=10, bold=True)
cover.cell(r, 3, "55").font = Font(name="Calibri", size=10, bold=True)
cover.cell(r, 3).alignment = Alignment(horizontal="center", vertical="top")
for c in (1, 2, 3):
    cover.cell(r, c).fill = LOC_FILL
    cover.cell(r, c).border = BORDER
cover.row_dimensions[r].height = 60
r += 2

cover.cell(r, 1, "Implementation phases (per the PRD)").font = Font(name="Calibri", size=12, bold=True, color="305496")
r += 1
phases = [
    ("v2.5.0", "Specs 02 (schema only), 07 (settings + migration + tests). 1,000 LOC."),
    ("v2.5.0 part 2", "Specs 03 (runtime), 04 (scheduler + evaluator). 1,700 LOC."),
    ("v2.5.1", "Specs 01 (commands + tools), 05 (notifications + TUI), 06 (pi-subagents bridge). 1,500 LOC. 1,900 LOC of tests across both releases."),
]
for phase, desc in phases:
    cover.cell(r, 1, phase).font = Font(name="Calibri", size=10, bold=True)
    cover.cell(r, 1).alignment = WRAP
    cover.cell(r, 2, desc).alignment = WRAP
    cover.cell(r, 2).font = BODY_FONT
    cover.row_dimensions[r].height = 50
    r += 1

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  sheets: {wb.sheetnames}")
total = 0
for name, rows in SHEETS:
    print(f"  {name}: {len(rows)} requirements")
    total += len(rows)
print(f"  total: {total} requirements")
