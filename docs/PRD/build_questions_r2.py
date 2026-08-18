"""
Build the sub-agent clarifying-questions round-2 xlsx.

4 follow-up questions covering the new constraints that came out of
round 1:
  R2-1  /loop command syntax for sub-agent mode (D3 follow-up)
  R2-2  Default tools for sub-agent loops — D7 revisited
  R2-3  Built-in roles (the pi-subagents analogy)
  R2-4  Goal / role / success / failure criteria + optional state file (D13 expansion)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\dev\pi-loop\wt-pi-subagent\docs\PRD\sub-agent-questions-r2.xlsx"

QUESTIONS = [
    # -------- R2-1: /loop syntax -----------------------------------------
    (
        "R2-1",
        "/loop command syntax",
        "How does a user opt a loop into sub-agent mode from the /loop command line?",
        "A) Add a flag: `/loop <interval> <prompt> --isolation sub-agent` (and other flags as needed)\n"
        "B) Move to a TUI form for sub-agent loops; keep the inline `<interval> <prompt>` for in-process only\n"
        "C) Add a long-form: `/loop sub-agent <interval> <prompt>` as a parallel sub-command",
        "A — add flags",
        "A — add flags (recommended for v2.5.0)",
        "A flag is the smallest change to the existing command: `LoopCreate` already takes a structured input; the slash command is a thin wrapper. The other fields (--role, --goal, --success-criteria, --failure-criteria, --state-file) become more flags. TUI form is great for sub-agent loops that need many fields, but the inline form should still work for the common case.",
        "B is a UX regression for power users who like the inline form; C duplicates the parsing logic.",
    ),
    # -------- R2-2: Default tools (D7 revisit) ---------------------------
    (
        "R2-2",
        "Default tools for sub-agent loops",
        "Given the new role concept (R2-3) and your note on D7 ('loops should have access to all tools/skills just as the main shell'), what is the default tool surface for a sub-agent loop iteration?",
        "A) Full surface (all tools the parent has, plus all skills). The user opts out by setting `subAgent.tools` per-loop.\n"
        "B) Read-only by default. The user opts in to write/edit/bash per-loop.\n"
        "C) Role-driven (the default depends on the loop's `role`; e.g. a `scout` loop is read-only, a `worker` loop has full tools). The user overrides per-loop with `subAgent.tools`.",
        "C — role-driven",
        "C — role-driven (recommended)",
        "Roles are the right place to encode 'what kind of work is this' — a scout is read-only by definition, a worker writes. Making the toolset role-driven keeps each role's surface sensible without per-loop configuration. Per-loop override (`subAgent.tools`) still wins, so power users can tweak.",
        "A is the user's literal D7 ask but loses the safety net; B is the original recommendation but adds friction for the common case (a worker loop that wants to write); C is the clean answer once roles exist.",
    ),
    # -------- R2-3: Built-in roles ---------------------------------------
    (
        "R2-3",
        "Built-in roles",
        "What default roles ship with pi-loop's sub-agent feature? (Pi-subagents ships scout/researcher/worker/reviewer/oracle/delegate; should we mirror any of those?)",
        "A) None — every sub-agent loop is a generic prompt. The user writes the system-prompt role inline.\n"
        "B) Mirror pi-subagents (scout, researcher, worker, reviewer, oracle) so users moving between the two packages have a familiar vocabulary.\n"
        "C) A minimal starter set: `scout` (read-only), `worker` (read+write), `digest-writer` (read+write-bounded). Three roles that cover the most common sub-agent use cases.",
        "C — minimal starter set",
        "C — minimal starter set (recommended)",
        "Three roles is the smallest set that demonstrates the pattern without bloating the extension. `scout` covers the read-only case; `worker` covers the write-capable case; `digest-writer` is a `worker` variant for the use case the PRD explicitly targets (your weekly digester loop). Users who want pi-subagents-style roles can install pi-subagents and the role vocabulary merges automatically via the background-work bridge.",
        "A forces every user to write their own system prompt; B bundles roles the user may not need and risks drift from pi-subagents' own definitions.",
    ),
    # -------- R2-4: Goal / role / success / failure + state file --------
    (
        "R2-4",
        "Goal / success / failure criteria + state file",
        "You wrote on D13: 'loops that are created must have a defined goal, role, success criteria, failure criteria. optionally a state management file or somewhere to look at.' Confirm the schema and the per-field requirement level.",
        "A) Optional on every loop. The TUI / cyclic form prompts for them but lets the user skip; the runtime is permissive.\n"
        "B) Required for sub-agent loops (the cyclic form refuses to save without them); optional for in-process loops.\n"
        "C) Required for ALL loops (in-process and sub-agent). A step backward from v2.x permissiveness but a step up in discipline.",
        "B — required for sub-agent only",
        "B — required for sub-agent only (recommended)",
        "Sub-agent loops are unsupervised — they run without the user watching — so they need explicit success/failure criteria to know when to pause, escalate, or back off. In-process loops have the user as the implicit success criterion (the user reads the wake and acts). Required-for-all (C) breaks the v2.x onboarding flow for users who just want a simple polling loop; optional (A) defeats the purpose of the field.",
        "The criteria fields are: `goal: string` (the loop's purpose), `successCriteria: string` (when an iteration counts as success), `failureCriteria: string` (when an iteration counts as failure, optionally with retry/backoff). `stateFile: string` (optional) points at a JSON file the child reads/writes for cross-iteration state. The criteria are evaluated by the parent against the child's `result.md` (or the LLM in the parent) after each iteration; the state file is a simple key/value JSON document the child manages directly.",
    ),
]

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Decision Matrix R2"

headers = [
    "ID",
    "Topic",
    "Decision / Question",
    "Options",
    "Recommendation",
    "Your choice (pre-filled with default — overwrite to change)",
    "Rationale",
    "Impact if changed",
]
ws.append(headers)

# Header style
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, _ in enumerate(headers, 1):
    cell = ws.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Body style
body_font = Font(name="Calibri", size=10, color="000000")
input_font = Font(name="Calibri", size=10, color="0000FF")
wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
recommend_fill = PatternFill("solid", fgColor="FFF2CC")
for q in QUESTIONS:
    qid, topic, question, options, recommend, your_choice, rationale, impact = q
    row = [qid, topic, question, options, recommend, your_choice, rationale, impact]
    ws.append(row)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(r, c)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = border
    ws.cell(r, 6).font = input_font
    ws.cell(r, 5).fill = recommend_fill

widths = {
    1: 8,    # ID
    2: 30,   # Topic
    3: 55,   # Decision / Question
    4: 60,   # Options
    5: 32,   # Recommendation
    6: 55,   # Your choice
    7: 55,   # Rationale
    8: 55,   # Impact
}
for c, w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 220

ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = "70AD47"  # green to distinguish from R1

# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------
legend = wb.create_sheet("How to answer")
legend["A1"] = "Sub-agent PRD — Round 2 Decision Matrix"
legend["A1"].font = Font(name="Calibri", size=14, bold=True, color="305496")
legend["A3"] = "What this is"
legend["A3"].font = Font(name="Calibri", size=11, bold=True)
legend["B3"] = ("4 follow-up questions raised by your answers to round 1. Each one ties to a "
                "specific round-1 answer (D3, D7, D13) and tightens the design. The format is "
                "the same as round 1: 4 columns of context, your choice in column F.")
legend["B3"].alignment = wrap

legend["A6"] = "Tie-back to round 1"
legend["A6"].font = Font(name="Calibri", size=11, bold=True)
legend["B6"] = (
    "R2-1  ← D3  You asked: what does the /loop syntax look like for sub-agent mode?\n"
    "R2-2  ← D7  You wrote: loops should have full tools; non-loop subagents can have different roles\n"
    "R2-3  ← D7  You mentioned: 'different types of subagents similar to the pi-subagents package'\n"
    "R2-4  ← D13 You added: every loop needs goal/role/success/failure + optional state file"
)
legend["B6"].alignment = wrap

legend["A11"] = "How to answer"
legend["A11"].font = Font(name="Calibri", size=11, bold=True)
legend["B11"] = (
    "1. Open 'Decision Matrix R2'.\n"
    "2. For each row, type your choice in column F (blue).\n"
    "3. If you want a hybrid, write it in plain English.\n"
    "4. Save and send back. If everything is answerable, I will update the PRD "
    "(docs/PRD/sub-agent.md) and start Phase 1 (v2.5.0) implementation.\n"
    "5. If a third round is needed, it will be small (1-2 questions at most)."
)
legend["B11"].alignment = wrap

legend["A18"] = "After round 2"
legend["A18"].font = Font(name="Calibri", size=11, bold=True)
legend["B18"] = (
    "The PRD will be amended to add: \n"
    "  - `goal`, `role`, `successCriteria`, `failureCriteria`, `stateFile` on LoopEntry (from R2-4)\n"
    "  - Three default roles: `scout`, `worker`, `digest-writer` (from R2-3)\n"
    "  - Role-driven tool allowlists (from R2-2)\n"
    "  - `--isolation`, `--role`, `--goal`, `--success-criteria`, `--failure-criteria`, `--state-file` flags on /loop (from R2-1)\n"
    "  - A goal/success/failure evaluator in the parent (likely a small LLM call after each sub-agent iteration, or a regex match on the result.md)\n\n"
    "Roughly +400 LOC over the original v2.5.0 estimate of 3,900. Total target stays in the 4,000-4,500 range."
)
legend["B18"].alignment = wrap

legend.column_dimensions["A"].width = 22
legend.column_dimensions["B"].width = 100
for r in (3, 6, 11, 18):
    legend.row_dimensions[r].height = 130
legend.row_dimensions[1].height = 24

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  sheets: {wb.sheetnames}")
print(f"  questions: {len(QUESTIONS)}")
